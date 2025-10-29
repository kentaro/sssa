#!/usr/bin/env python3
"""
宇宙スキル標準ExcelファイルからSkill Levelsを正しく抽出するスクリプト
"""

import openpyxl
import yaml
import sys

def extract_skill_levels(excel_file):
    """スキルレベル一覧を抽出"""
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb['④‐2スキルレベル一覧']

    skill_levels = []

    # データは行5から開始（行1-4はヘッダー）
    row_idx = 5
    while row_idx <= sheet.max_row:
        # カテゴリ（列B）
        category_cell = sheet.cell(row=row_idx, column=2)
        category = category_cell.value if category_cell.value else ''
        category = category.strip() if isinstance(category, str) else ''

        # スキル番号（列D）
        skill_number_cell = sheet.cell(row=row_idx, column=4)
        skill_number = skill_number_cell.value

        # スキル名（列E）
        skill_name_cell = sheet.cell(row=row_idx, column=5)
        skill_name = skill_name_cell.value if skill_name_cell.value else ''
        skill_name = skill_name.strip() if isinstance(skill_name, str) else ''

        # 空行ならスキップ
        if not skill_number or not skill_name:
            row_idx += 1
            continue

        # このスキルの4つの評価軸を読み込む
        for axis_offset in range(4):
            current_row = row_idx + axis_offset

            # 評価軸名（列F）
            eval_axis_cell = sheet.cell(row=current_row, column=6)
            eval_axis = eval_axis_cell.value if eval_axis_cell.value else ''
            eval_axis = eval_axis.strip() if isinstance(eval_axis, str) else ''

            if not eval_axis:
                continue

            # レベル1-5の説明（列G-K）
            levels = {}
            for level in range(1, 6):
                col_idx = 6 + level  # G=7, H=8, I=9, J=10, K=11
                level_cell = sheet.cell(row=current_row, column=col_idx)
                level_desc = level_cell.value if level_cell.value else ''
                level_desc = level_desc.strip() if isinstance(level_desc, str) else ''

                # 空の場合は'ー'として扱う
                if not level_desc:
                    level_desc = 'ー'

                levels[level] = level_desc

            # スキルレベル情報を追加
            skill_level = {
                'category': category,
                'skill_number': int(skill_number) if isinstance(skill_number, (int, float)) else skill_number,
                'skill_name': skill_name,
                'evaluation_axis': eval_axis,
                'levels': levels
            }
            skill_levels.append(skill_level)

        # 次のスキル（4行スキップ）
        row_idx += 4

    return skill_levels

def main():
    excel_file = sys.argv[1] if len(sys.argv) > 1 else 'uchuskill2025.xlsx'
    output_file = sys.argv[2] if len(sys.argv) > 2 else 'skill_levels_fixed.yaml'

    print(f"Excelファイルを読み込み中: {excel_file}")
    skill_levels = extract_skill_levels(excel_file)

    print(f"抽出したスキルレベル: {len(skill_levels)}件")

    # 最初の数件を確認
    print("\n最初の2件:")
    for i, sl in enumerate(skill_levels[:2]):
        print(f"\n{i+1}.")
        print(f"  スキル番号: {sl['skill_number']}")
        print(f"  スキル名: {sl['skill_name']}")
        print(f"  評価軸: {sl['evaluation_axis']}")
        print(f"  レベル数: {len(sl['levels'])}")
        if sl['levels']:
            level1 = sl['levels'][1]
            print(f"  レベル1: {level1[:50]}...")

    # YAMLファイルに保存
    output_data = {'skill_levels': skill_levels}
    with open(output_file, 'w', encoding='utf-8') as f:
        yaml.dump(output_data, f, allow_unicode=True, sort_keys=False, default_flow_style=False)

    print(f"\n✅ 保存完了: {output_file}")

if __name__ == '__main__':
    main()
