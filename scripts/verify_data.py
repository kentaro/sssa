#!/usr/bin/env python3
"""
宇宙スキル標準データの整合性検証スクリプト

使用方法:
    python3 scripts/verify_data.py

このスクリプトは、元のExcelファイルと抽出したYAMLファイルの整合性を検証します。
"""

import sys
import os
from pathlib import Path

try:
    import openpyxl
    import yaml
except ImportError:
    print("エラー: 必要なライブラリがインストールされていません。")
    print("以下のコマンドでインストールしてください:")
    print("  pip install openpyxl pyyaml")
    sys.exit(1)


class DataVerifier:
    def __init__(self, excel_path, yaml_path):
        self.excel_path = excel_path
        self.yaml_path = yaml_path
        self.errors = []
        self.warnings = []

    def load_data(self):
        """データを読み込む"""
        print("データを読み込み中...")
        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)

        with open(self.yaml_path, 'r', encoding='utf-8') as f:
            self.yaml_data = yaml.safe_load(f)

        print("✓ データ読み込み完了\n")

    def verify_counts(self):
        """件数の検証"""
        print("=" * 80)
        print("【1. 件数検証】")
        print("=" * 80)

        checks = [
            ('スキル', '①スキル一覧', 4, 3, 'skills'),
            ('業務', '②業務一覧', 4, 3, 'tasks'),
            ('スキルディクショナリ', '③スキルディクショナリ', 4, 3, 'skill_dictionary'),
        ]

        for name, sheet_name, min_row, col_idx, yaml_key in checks:
            excel_count = 0
            sheet = self.wb[sheet_name]
            for row in sheet.iter_rows(min_row=min_row, values_only=True):
                if row[col_idx] is not None:
                    excel_count += 1

            yaml_count = len(self.yaml_data[yaml_key])

            match = excel_count == yaml_count
            status = "✓" if match else "✗"

            print(f"{name}:")
            print(f"  Excel: {excel_count:3d}件")
            print(f"  YAML:  {yaml_count:3d}件")
            print(f"  結果:  {status}")

            if not match:
                self.errors.append(f"{name}の件数が一致しません (Excel: {excel_count}, YAML: {yaml_count})")
            print()

        # ロールは別途カウント（プレースホルダーを除外）
        name = 'ロール'
        sheet_name = '⑥ロール一覧'
        yaml_key = 'roles'

        excel_count = 0
        sheet = self.wb[sheet_name]
        for row in sheet.iter_rows(min_row=4, values_only=True):
            col3 = row[2]  # #（番号）
            col4 = row[3]  # ロール（小）
            if col3 and col4 and col4 != '*':
                excel_count += 1

        yaml_roles = [r for r in self.yaml_data[yaml_key] if r.get('name') != '*']
        yaml_count = len(yaml_roles)

        match = excel_count == yaml_count
        status = "✓" if match else "✗"

        print(f"{name}:")
        print(f"  Excel: {excel_count:3d}件")
        print(f"  YAML:  {yaml_count:3d}件")
        print(f"  結果:  {status}")

        if not match:
            self.errors.append(f"{name}の件数が一致しません (Excel: {excel_count}, YAML: {yaml_count})")
        print()

    def verify_skills(self):
        """スキルデータの詳細検証"""
        print("=" * 80)
        print("【2. スキルデータ検証】")
        print("=" * 80)

        sheet = self.wb['①スキル一覧']
        excel_skills = []
        current_category = ""

        for row in sheet.iter_rows(min_row=4, values_only=True):
            if row[3] is not None:
                if row[1]:
                    current_category = row[1]
                excel_skills.append({
                    'category': current_category,
                    'number': row[3],
                    'name': row[4] if row[4] else "",
                    'description': row[5] if len(row) > 5 and row[5] else ""
                })

        # 全スキルを照合
        mismatch_count = 0
        for i, (excel_skill, yaml_skill) in enumerate(zip(excel_skills, self.yaml_data['skills'])):
            if excel_skill['number'] != yaml_skill['number']:
                self.errors.append(f"スキル#{i+1}: 番号不一致 (Excel: {excel_skill['number']}, YAML: {yaml_skill['number']})")
                mismatch_count += 1

            if excel_skill['name'] != yaml_skill['name']:
                self.errors.append(f"スキル#{excel_skill['number']}: 名前不一致")
                print(f"  Excel: {excel_skill['name']}")
                print(f"  YAML:  {yaml_skill['name']}")
                mismatch_count += 1

        if mismatch_count == 0:
            print("✓ 全95スキルの番号・名前が一致しています")
        else:
            print(f"✗ {mismatch_count}件の不一致が見つかりました")
        print()

    def verify_tasks(self):
        """業務データの詳細検証"""
        print("=" * 80)
        print("【3. 業務データ検証】")
        print("=" * 80)

        sheet = self.wb['②業務一覧']
        excel_tasks = []
        current_category = ""
        current_subcategory = ""

        for row in sheet.iter_rows(min_row=4, values_only=True):
            if row[3] is not None:
                if row[1]:
                    current_category = row[1]
                if row[2]:
                    current_subcategory = row[2]
                else:
                    current_subcategory = ""
                excel_tasks.append({
                    'category': current_category,
                    'subcategory': current_subcategory,
                    'number': row[3],
                    'name': row[4] if row[4] else ""
                })

        mismatch_count = 0
        for i, (excel_task, yaml_task) in enumerate(zip(excel_tasks, self.yaml_data['tasks'])):
            if excel_task['number'] != yaml_task['number']:
                self.errors.append(f"業務#{i+1}: 番号不一致")
                mismatch_count += 1

            if excel_task['name'] != yaml_task['name']:
                self.errors.append(f"業務#{excel_task['number']}: 名前不一致")
                mismatch_count += 1

        if mismatch_count == 0:
            print("✓ 全58業務の番号・名前が一致しています")
        else:
            print(f"✗ {mismatch_count}件の不一致が見つかりました")
        print()

    def verify_skill_levels(self):
        """スキルレベル定義の検証"""
        print("=" * 80)
        print("【4. スキルレベル定義検証】")
        print("=" * 80)

        sheet = self.wb['④‐2スキルレベル一覧']
        excel_count = 0

        # Excelのスキルレベル定義を正しくカウント
        # 各スキルは4行（4つの評価軸）を持つ
        row_idx = 5
        while row_idx <= sheet.max_row:
            skill_number = sheet.cell(row=row_idx, column=4).value
            if not skill_number:
                row_idx += 1
                continue

            # このスキルの4つの評価軸をカウント
            for axis_offset in range(4):
                current_row = row_idx + axis_offset
                eval_axis = sheet.cell(row=current_row, column=6).value
                if eval_axis:
                    excel_count += 1

            # 次のスキル（4行スキップ）
            row_idx += 4

        yaml_count = len(self.yaml_data['skill_levels'])

        print(f"スキルレベル定義数:")
        print(f"  Excel: {excel_count}件")
        print(f"  YAML:  {yaml_count}件")

        if excel_count == yaml_count:
            print(f"  結果:  ✓")
        else:
            print(f"  結果:  ✗")
            self.errors.append(f"スキルレベル定義数が異なります (差分: {abs(excel_count - yaml_count)}件)")

        # サンプルチェック（最初のスキルレベル定義）
        if self.yaml_data['skill_levels']:
            sample = self.yaml_data['skill_levels'][0]
            print(f"\nサンプル確認:")
            print(f"  スキル番号: {sample['skill_number']}")
            print(f"  スキル名: {sample['skill_name']}")
            print(f"  評価軸: {sample['evaluation_axis']}")
            print(f"  レベル数: {len(sample['levels'])}段階")

            # データ構造の検証
            if not isinstance(sample['skill_number'], int):
                self.errors.append("skill_numberが整数型ではありません")
            if len(sample['levels']) != 5:
                self.warnings.append("一部のスキルレベルが5段階ではありません")

            # レベル1の説明が評価軸名になっていないか確認（以前のバグ）
            level_1_desc = sample['levels'].get(1, '')
            if level_1_desc in ['遂行可能な業務範囲・深さ', '業務遂行時の自立性', '資格・検定', '経験年数']:
                self.errors.append("skill_levelsのデータ構造が不正です（フィールドがずれています）")
        print()

    def verify_roles(self):
        """ロールデータの詳細検証"""
        print("=" * 80)
        print("【5. ロールデータ検証】")
        print("=" * 80)

        sheet = self.wb['⑥ロール一覧']
        excel_roles = []
        current_category = ''

        for row in sheet.iter_rows(min_row=4, values_only=True):
            col2 = row[1]  # ロール（大）
            col3 = row[2]  # #（番号）
            col4 = row[3]  # ロール（小）

            if col3 and col4 and col4 != '*':
                if col2:
                    current_category = col2
                excel_roles.append({
                    'category': current_category,
                    'number': col3,
                    'name': col4
                })

        yaml_roles = [r for r in self.yaml_data['roles'] if r.get('name') != '*']

        print(f"ロール数:")
        print(f"  Excel: {len(excel_roles)}件")
        print(f"  YAML:  {len(yaml_roles)}件")

        if len(excel_roles) == len(yaml_roles):
            print(f"  結果:  ✓")
        else:
            print(f"  結果:  ✗")
            self.errors.append(f"ロール数が一致しません (Excel: {len(excel_roles)}, YAML: {len(yaml_roles)})")

        # カテゴリ別の集計
        excel_cat_count = {}
        yaml_cat_count = {}

        for role in excel_roles:
            cat = role['category'] if role['category'] else '（カテゴリなし）'
            excel_cat_count[cat] = excel_cat_count.get(cat, 0) + 1

        for role in yaml_roles:
            cat = role['category'] if role['category'] else '（カテゴリなし）'
            yaml_cat_count[cat] = yaml_cat_count.get(cat, 0) + 1

        print(f"\nカテゴリ別ロール数:")
        all_cats = set(excel_cat_count.keys()) | set(yaml_cat_count.keys())
        all_match = True

        for cat in sorted(all_cats):
            excel_count = excel_cat_count.get(cat, 0)
            yaml_count = yaml_cat_count.get(cat, 0)
            match = excel_count == yaml_count
            status = "✓" if match else "✗"

            if not match:
                all_match = False
                self.errors.append(f"ロールカテゴリ「{cat}」の件数が一致しません")

            cat_display = cat.replace('\n', ' ')[:40]
            print(f"  {cat_display:40s}: Excel={excel_count:2d}, YAML={yaml_count:2d} {status}")

        if all_match and len(excel_roles) == len(yaml_roles):
            print(f"\n✓ 全ロールの件数とカテゴリが一致しています")
        else:
            print(f"\n✗ 一部のロールで不一致があります")
        print()

    def verify_categories(self):
        """カテゴリ分類の検証"""
        print("=" * 80)
        print("【6. カテゴリ分類検証】")
        print("=" * 80)

        # スキルカテゴリ
        excel_categories = {}
        sheet = self.wb['①スキル一覧']
        current_category = ""

        for row in sheet.iter_rows(min_row=4, values_only=True):
            if row[3] is not None:
                if row[1]:
                    current_category = row[1]
                if current_category not in excel_categories:
                    excel_categories[current_category] = 0
                excel_categories[current_category] += 1

        yaml_categories = {}
        for skill in self.yaml_data['skills']:
            cat = skill['category']
            if cat not in yaml_categories:
                yaml_categories[cat] = 0
            yaml_categories[cat] += 1

        print("スキルカテゴリ別件数:")
        all_match = True
        for cat in sorted(excel_categories.keys()):
            excel_count = excel_categories.get(cat, 0)
            yaml_count = yaml_categories.get(cat, 0)
            match = excel_count == yaml_count
            status = "✓" if match else "✗"

            if not match:
                all_match = False
                self.errors.append(f"カテゴリ「{cat}」の件数が一致しません")

            # カテゴリ名を表示用に整形（改行を除去）
            cat_display = cat.replace('\n', ' ')[:30]
            print(f"  {cat_display:30s}: Excel={excel_count:2d}, YAML={yaml_count:2d} {status}")

        if all_match:
            print(f"\n✓ 全カテゴリの件数が一致しています")
        else:
            print(f"\n✗ 一部のカテゴリで不一致があります")
        print()

    def print_summary(self):
        """検証結果のサマリーを表示"""
        print("=" * 80)
        print("【検証結果サマリー】")
        print("=" * 80)

        if not self.errors and not self.warnings:
            print("✓ すべての検証項目に合格しました！")
            print("  Excel と YAML のデータは完全に整合しています。")
        else:
            if self.errors:
                print(f"✗ {len(self.errors)}件のエラーが見つかりました:")
                for error in self.errors:
                    print(f"  - {error}")

            if self.warnings:
                print(f"\n⚠ {len(self.warnings)}件の警告があります:")
                for warning in self.warnings:
                    print(f"  - {warning}")

        print("=" * 80)

        return len(self.errors) == 0

    def run(self):
        """検証を実行"""
        self.load_data()
        self.verify_counts()
        self.verify_skills()
        self.verify_tasks()
        self.verify_skill_levels()
        self.verify_roles()
        self.verify_categories()

        return self.print_summary()


def main():
    # パスの設定
    script_dir = Path(__file__).parent
    project_root = script_dir.parent

    excel_path = project_root / "tmp" / "uchuskill2025.xlsx"
    yaml_path = project_root / "data" / "space_skill_standard_complete.yaml"

    # Excelファイルが存在しない場合は/tmpから探す
    if not excel_path.exists():
        excel_path = Path("/tmp/uchuskill2025.xlsx")

    if not excel_path.exists():
        print("エラー: Excelファイルが見つかりません")
        print(f"パス: {excel_path}")
        print("\n以下のコマンドでダウンロードしてください:")
        print('  curl -L -o /tmp/uchuskill2025.xlsx "https://www8.cao.go.jp/space/skill/uchuskill2025.xlsx"')
        sys.exit(1)

    if not yaml_path.exists():
        print(f"エラー: YAMLファイルが見つかりません: {yaml_path}")
        sys.exit(1)

    print("宇宙スキル標準データ整合性検証")
    print("=" * 80)
    print(f"Excel: {excel_path}")
    print(f"YAML:  {yaml_path}")
    print("=" * 80)
    print()

    verifier = DataVerifier(excel_path, yaml_path)
    success = verifier.run()

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
